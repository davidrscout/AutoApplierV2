from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from .utils import DATA_DIR


TRACKING_PATH = DATA_DIR / "applications.xlsx"
HEADERS = [
    "timestamp",
    "role",
    "company",
    "job_title",
    "job_url",
    "location",
    "remote",
    "cv_used",
    "match_score",
    "status",
    "notes",
]


def ensure_tracking_file() -> None:
    TRACKING_PATH.parent.mkdir(parents=True, exist_ok=True)
    if not TRACKING_PATH.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "applications"
        ws.append(HEADERS)
        wb.save(TRACKING_PATH)
        return
    wb = load_workbook(TRACKING_PATH)
    ws = wb.active
    existing = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if existing != HEADERS:
        new_title = "applications_v2"
        if new_title in wb.sheetnames:
            ws = wb[new_title]
        else:
            ws = wb.create_sheet(new_title)
            ws.append(HEADERS)
        wb.save(TRACKING_PATH)


def log_application(
    role: str,
    company: str,
    job_title: str,
    job_url: str,
    location: str,
    remote: str,
    cv_used: str,
    score: float | None,
    status: str,
    notes: str,
) -> None:
    ensure_tracking_file()
    wb = load_workbook(TRACKING_PATH)
    ws = wb.active
    timestamp = datetime.utcnow().isoformat()
    ws.append(
        [
            timestamp,
            role,
            company,
            job_title,
            job_url,
            location,
            remote,
            cv_used,
            score if score is not None else "",
            status,
            notes,
        ]
    )
    wb.save(TRACKING_PATH)
